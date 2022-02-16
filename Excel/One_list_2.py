import openpyxl as xl

wb = xl.load_workbook('trekking1.xlsx')
ws = wb.active

res_dict = dict()
for i in range(37):
    res_dict[ws.cell(row=i+2, column=1).value] = ws.cell(row=i+2,column=2).value
print(*sorted(res_dict, key=lambda x: (-res_dict[x], x)), sep="\n")
