from openpyxl import load_workbook, workbook

wb = load_workbook('salaries.xlsx')
ws = wb.active
# c = ws['12'].value


dict_cities = dict()
dict_occupations = dict()
for i in range(2, 10):
    dict_cities[i-2] = ws['A'+str(i)].value
    # print(ws['A'+str(i)].value)
# print(dict_cities)
for i in range(7):
    dict_occupations[i] = ws[chr(i+66)+'1'].value
    # print(ws['A'+str(i)].value)
# print(dict_occupations)


data = list()
row_data = list()
for i in range(2, 10):
    for j in range(7):
        # print(chr(j+66), str(i), ws[chr(j+66)+str(i)].value)
        row_data.append(ws[chr(j+66)+str(i)].value)
    data.append(row_data.copy())
    row_data.clear()
# print(data)

# print(list(map(lambda x: sorted(x), data)))
# print(list(map(sorted, data)))
# print(list(map(lambda x: sorted(x)[3], data)))
print(dict_cities[list(map(lambda x: sorted(x)[3], data)).index(max(map(lambda x: sorted(x)[3], data)))], end = ' ')


# print(data)
row_data.clear()
for i in range(len(dict_occupations)):
    # for j in range(len(dict_cities)):
    # print(sum(map(lambda x: x[i], data))/len(dict_cities))
    row_data.append(sum(map(lambda x: x[i], data))/len(dict_cities))
print(dict_occupations[row_data.index(max(row_data))])
