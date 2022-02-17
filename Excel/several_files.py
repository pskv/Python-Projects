import urllib.request
import zipfile
import io
import openpyxl as xl

res_dict = dict()
with urllib.request.urlopen('https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip') as f:
    input_zip = zipfile.ZipFile(io.BytesIO(f.read()))
    for name in input_zip.namelist():
        wb = xl.load_workbook(io.BytesIO(input_zip.read(name)))
        ws = wb.active
        res_dict[ws['B2'].value] = ws['D2'].value

with open('report.txt', 'w', encoding='utf-8') as output_f:
    for val in sorted(res_dict):
        output_f.write(val + ' ' + str(res_dict[val]) + '\n')
