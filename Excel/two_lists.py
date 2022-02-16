import openpyxl as xl

wb = xl.load_workbook('trekking2.xlsx')
ws = wb['Справочник']
dict1 = dict()
for i in range(37):
    dict1[ws.cell(row=i+2, column=1).value] = \
        tuple(ws.iter_rows(min_row=i+2, min_col=2, max_col=5, values_only=True))[0]
# print(dict1)
# print(*sorted(res_dict, key=lambda x: (-res_dict[x], x)), sep="\n")

res_list = list()
ws = wb['Раскладка']
for i in range(12):
    # print(ws['A'+str(i+2)].value, end=' ')
    # print(ws['B'+str(i+2)].value, end=' ')
    # print(dict1[ws['A' + str(i + 2)].value], end=' ')
    res_list.append(list(map(lambda x: ws['B'+str(i+2)].value * x / 100 if x is not None else 0,
                             dict1[ws['A' + str(i + 2)].value])))
# print(*res_list, sep="\n")

print(int(sum(map(lambda x: x[0], res_list))), end=' ')
print(int(sum(map(lambda x: x[1], res_list))), end=' ')
print(int(sum(map(lambda x: x[2], res_list))), end=' ')
print(int(sum(map(lambda x: x[3], res_list))), end=' ')
